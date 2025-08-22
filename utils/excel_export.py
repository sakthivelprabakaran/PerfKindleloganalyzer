import openpyxl
from openpyxl.styles import Font, PatternFill

class ExcelExporter:
    def __init__(self):
        pass

    def export_excel_with_highlighting(self, batch_results, filename):
        """Export to Excel with the new format."""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Batch Results"

            # Find max number of iterations for header
            max_iterations = 0
            for results in batch_results.values():
                if len(results) > max_iterations:
                    max_iterations = len(results)

            # Create headers
            headers = ["Test Case Name"]
            for i in range(1, max_iterations + 1):
                headers.append(f"IT_{i:02d}")
            headers.extend(["Average", "Waveform Data"])
            sheet.append(headers)

            # Write data rows
            for test_case_name, results in batch_results.items():
                row_data = [test_case_name]
                durations = [r['duration'] for r in results]

                # Add iteration durations
                for i in range(max_iterations):
                    if i < len(durations):
                        row_data.append(f"{durations[i]:.3f}")
                    else:
                        row_data.append("")

                # Add average
                avg_duration = sum(durations) / len(durations) if durations else 0
                row_data.append(f"{avg_duration:.3f}")

                # Add waveform data
                row_data.append(self.get_waveform_summary(results))

                sheet.append(row_data)

            # Auto-size columns
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(filename)
            return True, f"Excel file saved to:\n{filename}"

        except Exception as e:
            return False, f"Failed to save Excel file: {str(e)}"

    def generate_excel_bytes(self, batch_results):
        """Generate Excel report and return as bytes."""
        import io
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Batch Results"

            # ... (logic from export_excel_with_highlighting) ...
            max_iterations = 0
            for results in batch_results.values():
                if len(results) > max_iterations:
                    max_iterations = len(results)
            headers = ["Test Case Name"] + [f"IT_{i+1:02d}" for i in range(max_iterations)] + ["Average", "Waveform Data"]
            sheet.append(headers)
            for test_case_name, results in batch_results.items():
                row_data = [test_case_name]
                durations = [r['duration'] for r in results]
                for i in range(max_iterations):
                    row_data.append(f"{durations[i]:.3f}" if i < len(durations) else "")
                avg_duration = sum(durations) / len(durations) if durations else 0
                row_data.append(f"{avg_duration:.3f}")
                row_data.append(self.get_waveform_summary(results))
                sheet.append(row_data)

            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        except Exception as e:
            raise e

    def get_waveform_summary(self, results):
        """Get a summary of waveform data for a set of results."""
        if not results:
            return ""

        patterns = {}
        for result in results:
            waveform_data = []
            for idx, height_info in enumerate(result['all_heights'], 1):
                height = height_info['height']
                waveform = height_info['waveform']
                waveform_data.append(f"{idx}. Height - {height}, Waveform - {waveform}")

            pattern_key = "\n".join(waveform_data)
            if pattern_key not in patterns:
                patterns[pattern_key] = []
            patterns[pattern_key].append(f"IT_{result['iteration']:02d}")

        summary = []
        for pattern, iterations in patterns.items():
            if len(iterations) == len(results):
                summary.append("Same pattern for all iterations:\n" + pattern)
            else:
                summary.append(f"Pattern for {', '.join(iterations)}:\n" + pattern)

        return "\n\n".join(summary)
